"""
從「Xolair 客戶清單檢查.xlsx」匯出 seed_data.json，供 db/seed.js 匯入資料庫使用。

用法：
    python scripts/export_seed_data.py

需要先安裝 openpyxl（pip install openpyxl）。
"""
import json
import os

import openpyxl

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SOURCE_XLSX = os.path.join(BASE_DIR, 'Xolair 客戶清單檢查.xlsx')
OUTPUT_JSON = os.path.join(BASE_DIR, 'seed_data.json')


def main():
    wb = openpyxl.load_workbook(SOURCE_XLSX, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    customers = []
    psrs = {}
    for r in rows:
        if not r or not r[2]:
            continue
        specialty, tiering, psr_code, psr_name, cust_code, cust_name, contact_name, dept, title = r[:9]
        customers.append({
            'specialty': specialty, 'tiering': tiering, 'psr_code': psr_code, 'psr_name': psr_name,
            'customer_code': cust_code, 'customer_name': cust_name, 'contact_name': contact_name,
            'department': dept, 'title': title,
        })
        if psr_code:
            psrs[psr_code] = psr_name if psr_name and psr_name != psr_code else psrs.get(psr_code, psr_code)

    out = {
        'customers': customers,
        'psrs': [{'code': k, 'name': v} for k, v in sorted(psrs.items())],
    }
    with open(OUTPUT_JSON, 'w', encoding='utf-8') as f:
        json.dump(out, f, ensure_ascii=False, indent=2)

    print(f'customers: {len(customers)}, psrs: {len(psrs)} -> {OUTPUT_JSON}')


if __name__ == '__main__':
    main()
